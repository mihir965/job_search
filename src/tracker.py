"""Centralized Excel tracker operations - ALL Excel I/O goes through here."""
import logging
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime
from filelock import FileLock
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

TRACKER_FILE = Path(__file__).parent.parent / "outreach_tracker.xlsx"
LOCK_FILE = Path(__file__).parent.parent / "outreach_tracker.xlsx.lock"


class Tracker:
    """Manages all Excel tracker operations with file locking."""

    def __init__(self, tracker_path: Path = None):
        self.tracker_path = tracker_path or TRACKER_FILE
        self.lock = FileLock(str(self.tracker_path) + ".lock")

    # ── Initialization ──

    def create_tracker(self):
        """Create Excel tracker with all sheets and headers."""
        if self.tracker_path.exists():
            logger.warning(f"Tracker already exists at {self.tracker_path}")
            return

        logger.info("Creating new Excel tracker...")
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Create all sheets
        self._create_companies_sheet(wb)
        self._create_new_roles_sheet(wb)
        self._create_global_tracker_sheet(wb)
        self._create_contacts_sheet(wb)
        self._create_outreach_log_sheet(wb)
        self._create_email_drafts_sheet(wb)
        self._create_dashboard_sheet(wb)

        wb.save(self.tracker_path)
        logger.info(f"Created tracker at {self.tracker_path}")

    def _create_companies_sheet(self, wb: Workbook):
        """Create Companies sheet with headers and data validation."""
        ws = wb.create_sheet("Companies")

        # Headers
        headers = ["Company", "Tier", "Sector", "Careers URL", "Board Type",
                   "Sponsors H-1B", "Domain", "Notes"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Add data validation for dropdowns
        tier_dv = DataValidation(type="list", formula1='"Tier 1 - Dream,Tier 2 - Strong,Tier 3 - Backup"')
        sector_dv = DataValidation(type="list", formula1='"Quant/HFT,Systems/Infra,Fintech,FAANG,Startup,Other"')
        board_dv = DataValidation(type="list", formula1='"Greenhouse,Lever,Ashby,Workday,Custom"')
        h1b_dv = DataValidation(type="list", formula1='"Yes,No,Unknown"')

        ws.add_data_validation(tier_dv)
        ws.add_data_validation(sector_dv)
        ws.add_data_validation(board_dv)
        ws.add_data_validation(h1b_dv)

        tier_dv.add(f"B2:B1000")
        sector_dv.add(f"C2:C1000")
        board_dv.add(f"E2:E1000")
        h1b_dv.add(f"F2:F1000")

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['H'].width = 30

        # Add initial companies from CLAUDE.md spec
        companies_data = [
            ["Jane Street", "Tier 1", "Quant/HFT", "", "Custom", "Yes", "janestreet.com", ""],
            ["Citadel Securities", "Tier 1", "Quant/HFT", "", "Greenhouse", "Yes", "citadelsecurities.com", ""],
            ["Hudson River Trading", "Tier 1", "Quant/HFT", "", "Custom", "Yes", "hudsonrivertrading.com", ""],
            ["Two Sigma", "Tier 1", "Quant/HFT", "", "Greenhouse", "Yes", "twosigma.com", ""],
            ["Tower Research Capital", "Tier 1", "Quant/HFT", "", "Custom", "Yes", "tower-research.com", ""],
            ["Jump Trading", "Tier 1", "Quant/HFT", "", "Custom", "Yes", "jumptrading.com", ""],
            ["D.E. Shaw", "Tier 1", "Quant/HFT", "", "Custom", "Yes", "deshaw.com", ""],
            ["Five Rings", "Tier 1", "Quant/HFT", "", "Custom", "Yes", "fiverings.com", ""],
            ["Virtu Financial", "Tier 2", "Quant/HFT", "", "Greenhouse", "Yes", "virtu.com", ""],
            ["IMC Trading", "Tier 2", "Quant/HFT", "", "Greenhouse", "Yes", "imc.com", ""],
            ["DRW", "Tier 2", "Quant/HFT", "", "Custom", "Yes", "drw.com", ""],
            ["Optiver", "Tier 2", "Quant/HFT", "", "Greenhouse", "Yes", "optiver.com", ""],
            ["Wolverine Trading", "Tier 2", "Quant/HFT", "", "Custom", "Unknown", "wolve.com", ""],
            ["Akuna Capital", "Tier 2", "Quant/HFT", "", "Greenhouse", "Yes", "akunacapital.com", ""],
            ["Susquehanna (SIG)", "Tier 2", "Quant/HFT", "", "Custom", "Yes", "sig.com", ""],
            ["Cloudflare", "Tier 2", "Systems/Infra", "", "Greenhouse", "Yes", "cloudflare.com", ""],
            ["Databricks", "Tier 2", "Systems/Infra", "", "Greenhouse", "Yes", "databricks.com", ""],
        ]

        for row_data in companies_data:
            ws.append(row_data)

    def _create_new_roles_sheet(self, wb: Workbook):
        """Create New Roles sheet."""
        ws = wb.create_sheet("New Roles")

        headers = ["Date Found", "Company", "Role Title", "Location", "URL",
                   "Department", "Board Type", "Status", "HM Name", "HM Email",
                   "Outreach Sent?", "Notes"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data validation
        status_dv = DataValidation(type="list", formula1='"New,Reviewing,Applied,Skipped"')
        ws.add_data_validation(status_dv)
        status_dv.add("H2:H5000")

        outreach_dv = DataValidation(type="list", formula1='"Yes,No"')
        ws.add_data_validation(outreach_dv)
        outreach_dv.add("K2:K5000")

        # Column widths
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['E'].width = 50

    def _create_global_tracker_sheet(self, wb: Workbook):
        """Create Global Tracker sheet."""
        ws = wb.create_sheet("Global Tracker")

        headers = ["Date Found", "Company", "Role Title", "Location", "URL",
                   "Status", "Applied Date", "HM Name", "HM Email",
                   "Outreach Status", "Follow-up Due", "Notes"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(bold=True, color="000000")

        # Data validation
        status_dv = DataValidation(type="list",
                                   formula1='"New,Applied,Emailed HM,Phone Screen,Interview,Rejected,Offer,Skipped"')
        ws.add_data_validation(status_dv)
        status_dv.add("F2:F10000")

        outreach_dv = DataValidation(type="list", formula1='"Not Sent,Sent,Replied,Meeting Scheduled"')
        ws.add_data_validation(outreach_dv)
        outreach_dv.add("J2:J10000")

        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['E'].width = 50

    def _create_contacts_sheet(self, wb: Workbook):
        """Create Contacts sheet."""
        ws = wb.create_sheet("Contacts")

        headers = ["Company", "Name", "Title", "Email", "LinkedIn",
                   "Type", "Source", "Confidence", "Email Verified", "Date Found"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data validation
        type_dv = DataValidation(type="list", formula1='"Hiring Manager,Recruiter,Engineer,Other"')
        ws.add_data_validation(type_dv)
        type_dv.add("F2:F5000")

        confidence_dv = DataValidation(type="list", formula1='"High,Medium,Low"')
        ws.add_data_validation(confidence_dv)
        confidence_dv.add("H2:H5000")

        verified_dv = DataValidation(type="list", formula1='"Yes,No"')
        ws.add_data_validation(verified_dv)
        verified_dv.add("I2:I5000")

        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['E'].width = 40

    def _create_outreach_log_sheet(self, wb: Workbook):
        """Create Outreach Log sheet."""
        ws = wb.create_sheet("Outreach Log")

        headers = ["Date Sent", "Company", "Contact Name", "Contact Email",
                   "Email Type", "Subject Line", "Role Referenced",
                   "Status", "Follow-up Due", "LLM Generated?", "Notes"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data validation
        type_dv = DataValidation(type="list", formula1='"Cold HM,Cold Recruiter,Role-Specific,Follow-up"')
        ws.add_data_validation(type_dv)
        type_dv.add("E2:E5000")

        status_dv = DataValidation(type="list",
                                   formula1='"Sent,Opened,Replied,Meeting Scheduled,No Response,Bounced"')
        ws.add_data_validation(status_dv)
        status_dv.add("H2:H5000")

        llm_dv = DataValidation(type="list", formula1='"Yes,No"')
        ws.add_data_validation(llm_dv)
        llm_dv.add("J2:J5000")

        ws.column_dimensions['F'].width = 50

    def _create_email_drafts_sheet(self, wb: Workbook):
        """Create Email Drafts sheet for LLM-generated email review queue."""
        ws = wb.create_sheet("Email Drafts")

        headers = ["Date Created", "Company", "Contact Name", "Contact Email",
                   "Email Type", "Subject Line", "Body", "Role Referenced",
                   "Status", "Approved Date", "Notes"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data validation
        status_dv = DataValidation(
            type="list",
            formula1='"Pending Review,Approved,Rejected,Sent"'
        )
        ws.add_data_validation(status_dv)
        status_dv.add("I2:I5000")

        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 80

    def _create_dashboard_sheet(self, wb: Workbook):
        """Create Dashboard sheet with metrics."""
        ws = wb.create_sheet("Dashboard", 0)  # Make it first sheet

        # Title
        ws['A1'] = "Job Search Dashboard"
        ws['A1'].font = Font(size=18, bold=True)

        # Metrics
        metrics = [
            ("Total Companies Tracked", "=COUNTA(Companies!A:A)-1"),
            ("Total Roles Found", "=COUNTA('Global Tracker'!A:A)-1"),
            ("New Roles (This Week)", ""),  # TODO: Add formula
            ("Roles - Applied", "=COUNTIF('Global Tracker'!F:F,\"Applied\")"),
            ("Roles - Interviewing", "=COUNTIF('Global Tracker'!F:F,\"Interview\")+COUNTIF('Global Tracker'!F:F,\"Phone Screen\")"),
            ("Outreach Sent", "=COUNTA('Outreach Log'!A:A)-1"),
            ("Outreach - Replied", "=COUNTIF('Outreach Log'!H:H,\"Replied\")+COUNTIF('Outreach Log'!H:H,\"Meeting Scheduled\")"),
            ("Response Rate", ""),  # TODO: Add formula
        ]

        row = 3
        for label, formula in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            if formula:
                ws[f'B{row}'] = formula
                ws[f'B{row}'].font = Font(size=14)
            row += 1

    # ── Reading Operations ──

    def get_companies(self) -> List[Dict]:
        """Get all companies from Companies sheet."""
        with self.lock:
            if not self.tracker_path.exists():
                return []

            wb = load_workbook(self.tracker_path, read_only=True)
            ws = wb["Companies"]

            companies = []
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue

                company = dict(zip(headers, row))
                companies.append(company)

            wb.close()
            return companies

    def get_new_roles(self) -> List[Dict]:
        """Get all roles from New Roles sheet with status 'New'."""
        with self.lock:
            if not self.tracker_path.exists():
                return []

            wb = load_workbook(self.tracker_path, read_only=True)
            ws = wb["New Roles"]

            roles = []
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue

                role = dict(zip(headers, row))
                if role.get("Status") == "New":
                    roles.append(role)

            wb.close()
            return roles

    def get_contacts_needing_enrichment(self) -> List[str]:
        """Get list of company names that have a domain but no Hiring Manager contact yet."""
        with self.lock:
            if not self.tracker_path.exists():
                return []

            wb = load_workbook(self.tracker_path, read_only=True)

            # Build set of companies that already have an HM contact
            companies_with_hm: set = set()
            if "Contacts" in wb.sheetnames:
                ws_contacts = wb["Contacts"]
                for row in ws_contacts.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    company = row[0]
                    contact_type = row[5] if len(row) > 5 else None  # Type column
                    email = row[3] if len(row) > 3 else None  # Email column
                    if contact_type == "Hiring Manager" and email:
                        companies_with_hm.add(company.lower())

            # Return companies that have a domain but no HM contact
            ws = wb["Companies"]
            companies = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                company_name = row[0]
                domain = row[6] if len(row) > 6 else None

                if company_name and domain:
                    if company_name.lower() not in companies_with_hm:
                        companies.append(company_name)

            wb.close()
            return companies

    def get_pending_outreach(self) -> List[Dict]:
        """Get contacts who haven't been emailed yet.

        Cross-references Contacts sheet with Outreach Log to find contacts
        that have an email but haven't been sent outreach. Also checks
        New Roles for role-specific outreach where HM Email is populated.
        """
        with self.lock:
            if not self.tracker_path.exists():
                return []

            wb = load_workbook(self.tracker_path, read_only=True)

            # 1. Build set of already-emailed (email, role) pairs from Outreach Log
            emailed = set()
            ws_log = wb["Outreach Log"]
            for row in ws_log.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                contact_email = row[3]  # Contact Email
                role_ref = row[6] or ""  # Role Referenced
                if contact_email:
                    emailed.add((contact_email.lower(), role_ref.lower()))

            # Also exclude contacts with pending/approved drafts
            emailed_drafts = set()
            if "Email Drafts" in wb.sheetnames:
                ws_drafts = wb["Email Drafts"]
                for row in ws_drafts.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    status = row[8]  # Status
                    if status in ("Pending Review", "Approved"):
                        email = row[3]  # Contact Email
                        role = row[7] or ""  # Role Referenced
                        if email:
                            emailed_drafts.add((email.lower(), role.lower()))

            # 2. Check New Roles with HM Email populated but outreach not sent
            pending = []
            seen_keys = set()
            ws_roles = wb["New Roles"]
            for row in ws_roles.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                # Columns: Date Found, Company, Role Title, Location, URL,
                #          Department, Board Type, Status, HM Name, HM Email,
                #          Outreach Sent?, Notes
                status = row[7]
                hm_name = row[8]
                hm_email = row[9]
                outreach_sent = row[10]

                if not hm_email or outreach_sent == "Yes":
                    continue
                if status in ("Skipped",):
                    continue

                role_title = row[2] or ""
                key = (hm_email.lower(), role_title.lower())
                if key in emailed or key in emailed_drafts or key in seen_keys:
                    continue

                seen_keys.add(key)
                pending.append({
                    'name': hm_name or "",
                    'email': hm_email,
                    'company': row[1],
                    'role': role_title,
                    'role_url': row[4] or "",
                    'type': 'Hiring Manager',
                    'title': '',
                })

            # 3. Check Contacts sheet for contacts not yet emailed (generic outreach)
            ws_contacts = wb["Contacts"]
            contacts_headers = [cell.value for cell in ws_contacts[1]]
            for row in ws_contacts.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                contact = dict(zip(contacts_headers, row))
                email = contact.get("Email")
                if not email:
                    continue

                contact_type = contact.get("Type", "Other")
                if contact_type not in ("Hiring Manager", "Recruiter"):
                    continue

                key = (email.lower(), "")  # Generic outreach (no specific role)
                if key in emailed or key in emailed_drafts or key in seen_keys:
                    continue

                seen_keys.add(key)
                pending.append({
                    'name': contact.get("Name", ""),
                    'email': email,
                    'company': contact.get("Company", ""),
                    'role': None,
                    'role_url': '',
                    'type': contact_type,
                    'title': contact.get("Title", ""),
                })

            wb.close()
            return pending

    def get_due_followups(self) -> List[Dict]:
        """Get outreach entries that are due for follow-up."""
        # For now, return empty - we'll implement after outreach module
        return []

    # ── Writing Operations ──

    def add_company(self, name: str, tier: str = "Tier 3 - Backup",
                    sector: str = "Other", careers_url: str = "",
                    board_type: str = "Custom", sponsors: str = "Unknown",
                    domain: str = "", notes: str = "") -> bool:
        """Add a company to the Companies sheet if not already present. Returns True if added."""
        with self.lock:
            if not self.tracker_path.exists():
                return False

            wb = load_workbook(self.tracker_path)
            ws = wb["Companies"]

            # Check for duplicates
            name_lower = name.lower()
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and row[0].lower() == name_lower:
                    wb.close()
                    return False

            ws.append([name, tier, sector, careers_url, board_type,
                      sponsors, domain, notes])
            wb.save(self.tracker_path)
            logger.info(f"Added new company: {name}")
            return True

    def add_new_roles(self, roles: List[Dict]):
        """Append roles to New Roles sheet."""
        if not roles:
            return

        with self.lock:
            wb = load_workbook(self.tracker_path)
            ws = wb["New Roles"]

            for role in roles:
                ws.append([
                    role.get("date_found", datetime.now().isoformat()),
                    role.get("company"),
                    role.get("title"),
                    role.get("location"),
                    role.get("url"),
                    role.get("department", ""),
                    role.get("source", ""),
                    "New",  # Status
                    "",  # HM Name
                    "",  # HM Email
                    "No",  # Outreach Sent?
                    role.get("notes", ""),
                ])

            wb.save(self.tracker_path)
            logger.info(f"Added {len(roles)} new roles to tracker")

    def add_to_global_tracker(self, roles: List[Dict]):
        """Append roles to Global Tracker sheet."""
        if not roles:
            return

        with self.lock:
            wb = load_workbook(self.tracker_path)
            ws = wb["Global Tracker"]

            for role in roles:
                ws.append([
                    role.get("date_found", datetime.now().isoformat()),
                    role.get("company"),
                    role.get("title"),
                    role.get("location"),
                    role.get("url"),
                    "New",  # Status
                    "",  # Applied Date
                    "",  # HM Name
                    "",  # HM Email
                    "Not Sent",  # Outreach Status
                    "",  # Follow-up Due
                    role.get("notes", ""),
                ])

            wb.save(self.tracker_path)
            logger.info(f"Added {len(roles)} roles to global tracker")

    def add_contact(self, contact: Dict):
        """Add a contact to the Contacts sheet."""
        with self.lock:
            wb = load_workbook(self.tracker_path)
            ws = wb["Contacts"]

            ws.append([
                contact.get("company"),
                contact.get("name"),
                contact.get("title"),
                contact.get("email"),
                contact.get("linkedin", ""),
                contact.get("type", "Other"),
                contact.get("source"),
                contact.get("confidence", "Medium"),
                "Yes" if contact.get("email_verified") else "No",
                contact.get("date_found", datetime.now().isoformat()),
            ])

            wb.save(self.tracker_path)
            logger.debug(f"Added contact: {contact.get('name')} at {contact.get('company')}")

    def log_outreach(self, entry: Dict):
        """Log an outreach email to Outreach Log sheet."""
        with self.lock:
            wb = load_workbook(self.tracker_path)
            ws = wb["Outreach Log"]

            ws.append([
                entry.get("date_sent", datetime.now().isoformat()),
                entry.get("company"),
                entry.get("contact_name"),
                entry.get("contact_email"),
                entry.get("email_type"),
                entry.get("subject"),
                entry.get("role_referenced", ""),
                "Sent",  # Status
                entry.get("followup_due", ""),
                "Yes" if entry.get("llm_generated") else "No",
                entry.get("notes", ""),
            ])

            wb.save(self.tracker_path)
            logger.info(f"Logged outreach to {entry.get('contact_name')} at {entry.get('company')}")

    # ── Sheet Migration ──

    def _ensure_sheet_exists(self, sheet_name: str):
        """Ensure a sheet exists in the tracker, creating it if missing (for migration)."""
        with self.lock:
            if not self.tracker_path.exists():
                return

            wb = load_workbook(self.tracker_path)
            if sheet_name not in wb.sheetnames:
                create_method = f"_create_{sheet_name.lower().replace(' ', '_')}_sheet"
                if hasattr(self, create_method):
                    getattr(self, create_method)(wb)
                    wb.save(self.tracker_path)
                    logger.info(f"Migrated tracker: added '{sheet_name}' sheet")
                else:
                    logger.warning(f"No creation method for sheet '{sheet_name}'")
            wb.close()

    # ── Draft Queue Operations ──

    def save_draft(self, draft: Dict):
        """Save an LLM-generated email draft to the Email Drafts sheet."""
        self._ensure_sheet_exists("Email Drafts")

        with self.lock:
            wb = load_workbook(self.tracker_path)
            ws = wb["Email Drafts"]

            ws.append([
                draft.get("date_created", datetime.now().isoformat()),
                draft.get("company"),
                draft.get("contact_name"),
                draft.get("contact_email"),
                draft.get("email_type", "Role-Specific"),
                draft.get("subject"),
                draft.get("body"),
                draft.get("role_referenced", ""),
                "Pending Review",  # Status
                "",  # Approved Date
                draft.get("notes", "[LLM-generated]"),
            ])

            wb.save(self.tracker_path)
            logger.info(f"Saved draft for {draft.get('contact_name')} at {draft.get('company')}")

    def get_pending_drafts(self) -> List[Dict]:
        """Get all drafts with status 'Pending Review'."""
        return self._get_drafts_by_status("Pending Review")

    def get_approved_drafts(self) -> List[Dict]:
        """Get all drafts with status 'Approved'."""
        return self._get_drafts_by_status("Approved")

    def _get_drafts_by_status(self, status: str) -> List[Dict]:
        """Get drafts filtered by status, including their row number for updates."""
        self._ensure_sheet_exists("Email Drafts")

        with self.lock:
            if not self.tracker_path.exists():
                return []

            wb = load_workbook(self.tracker_path, read_only=True)
            ws = wb["Email Drafts"]

            headers = [cell.value for cell in ws[1]]
            drafts = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue

                draft = dict(zip(headers, row))
                if draft.get("Status") == status:
                    draft["_row"] = row_idx
                    drafts.append(draft)

            wb.close()
            return drafts

    def approve_draft(self, row: int):
        """Approve a draft by setting its status to 'Approved'."""
        self._update_draft_status(row, "Approved", approved_date=datetime.now().isoformat())

    def reject_draft(self, row: int):
        """Reject a draft by setting its status to 'Rejected'."""
        self._update_draft_status(row, "Rejected")

    def mark_draft_sent(self, row: int):
        """Mark a draft as sent."""
        self._update_draft_status(row, "Sent")

    def _update_draft_status(self, row: int, status: str, approved_date: str = ""):
        """Update a draft's status in-place."""
        self._ensure_sheet_exists("Email Drafts")

        with self.lock:
            wb = load_workbook(self.tracker_path)
            ws = wb["Email Drafts"]

            # Status is column I (9), Approved Date is column J (10)
            ws.cell(row=row, column=9, value=status)
            if approved_date:
                ws.cell(row=row, column=10, value=approved_date)

            wb.save(self.tracker_path)
            logger.debug(f"Updated draft row {row} to status '{status}'")


# Global singleton
tracker = Tracker()
